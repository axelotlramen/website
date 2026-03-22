import logging
import os
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scripts.constants import now

THREE_WEEKS = 21

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
INPUT_FONT = Font(color="0000FF", name="Arial")  # Blue = user-editable input
FORMULA_FONT = Font(color="000000", name="Arial")  # Black = formula
ALT_FILL = PatternFill("solid", start_color="D6E4F0")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


@dataclass
class GameConfig:
    name: str
    xlsx_file: str
    currency_name: str
    pull_item_name: str
    pull_cost: int
    five_star_pity: int
    diary_fetcher: callable  # type: ignore
    currency_attr: str


HSR_CONFIG = GameConfig(
    name="HSR",
    xlsx_file="data/hsr_diary_log.xlsx",
    currency_name="Stellar Jades",
    pull_item_name="Passes",
    pull_cost=160,
    five_star_pity=80,
    diary_fetcher=lambda client, uid: client.get_starrail_diary(uid=uid),
    currency_attr="current_hcoin",
)

GENSHIN_CONFIG = GameConfig(
    name="Genshin",
    xlsx_file="data/genshin_diary_log.xlsx",
    currency_name="Primogems",
    pull_item_name="Fates",
    pull_cost=160,
    five_star_pity=80,
    diary_fetcher=lambda client, uid: client.get_genshin_diary(uid=uid),
    currency_attr="current_primogems",
)

# Column layout (1-indexed):
# A=Date, B=Net Currency Gain, C=Pulls Net Gain, D=currency_name, E=Pulls,
# F=Total Pulls, G=Currency Needed for 5 Star, H=3-Week Avg Gain, I=Estimated Days Til 5 Star

COL_DATE = 1
COL_NET_CURRENCY = 2
COL_PULLS_NET = 3
COL_CURRENCY_TOTAL = 4
COL_PULLS_TOTAL = 5
COL_TOTAL_PULLS = 6
COL_CURRENCY_NEEDED = 7
COL_AVG_GAIN = 8
COL_EST_DAYS = 9


def _col(n):
    return get_column_letter(n)


def _apply_header(ws, currency_name):
    headers = [
        "Date",
        "Net Currency Gain",
        "Pulls Net Gain",
        currency_name,
        "Pulls",
        "Total Pulls",
        "Currency Needed for 5 Star",
        "3-Week Avg Gain",
        "Estimated Days Til 5 Star",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    col_widths = [12, 20, 16, 18, 8, 14, 28, 18, 26]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[_col(col)].width = width


def _write_row_formulas(ws, row, pull_cost, five_star_pity):
    """
    Write Excel formulas for a given data row.
    Columns B (Net Currency Gain) and C (Pulls Net Gain) are left as plain
    values (blue, user-editable inputs). All other calculated columns use
    Excel formulas that auto-update when B or C is changed.
    """
    r = row
    prev_r = r - 1  # previous data row (or 0 if this is the first)

    # D: currency_name total = previous total + Net Currency Gain
    if prev_r < 2:
        ws.cell(r, COL_CURRENCY_TOTAL).value = f"={_col(COL_NET_CURRENCY)}{r}"
    else:
        ws.cell(r, COL_CURRENCY_TOTAL).value = (
            f"={_col(COL_CURRENCY_TOTAL)}{prev_r}+{_col(COL_NET_CURRENCY)}{r}"
        )

    # E: Pulls total = previous pulls total + Pulls Net Gain
    if prev_r < 2:
        ws.cell(r, COL_PULLS_TOTAL).value = f"={_col(COL_PULLS_NET)}{r}"
    else:
        ws.cell(r, COL_PULLS_TOTAL).value = (
            f"={_col(COL_PULLS_TOTAL)}{prev_r}+{_col(COL_PULLS_NET)}{r}"
        )

    # F: Total Pulls = D/pull_cost + E
    ws.cell(r, COL_TOTAL_PULLS).value = (
        f"={_col(COL_CURRENCY_TOTAL)}{r}/{pull_cost}+{_col(COL_PULLS_TOTAL)}{r}"
    )

    # G: Currency Needed = MAX((pity - Total Pulls) * pull_cost, 0)
    ws.cell(r, COL_CURRENCY_NEEDED).value = (
        f"=MAX(({five_star_pity}-{_col(COL_TOTAL_PULLS)}{r})*{pull_cost},0)"
    )

    # H: 3-Week Avg Gain (average of last THREE_WEEKS combined daily gains)
    # Combined daily gain for each row = Net Currency Gain + Pulls Net Gain * pull_cost
    # We build a helper column approach using SUMPRODUCT over the last 21 rows.
    start_r = max(2, r - THREE_WEEKS + 1)
    b_range = f"{_col(COL_NET_CURRENCY)}{start_r}:{_col(COL_NET_CURRENCY)}{r}"
    c_range = f"{_col(COL_PULLS_NET)}{start_r}:{_col(COL_PULLS_NET)}{r}"
    count = r - start_r + 1
    ws.cell(r, COL_AVG_GAIN).value = (
        f"=IF({count}>={THREE_WEEKS},"
        f"SUMPRODUCT({b_range}+{c_range}*{pull_cost})/{THREE_WEEKS},"
        f"0)"
    )

    # I: Estimated Days = Currency Needed / Avg Gain (guarded against div/0)
    ws.cell(r, COL_EST_DAYS).value = (
        f"=IF({_col(COL_AVG_GAIN)}{r}>0,"
        f"{_col(COL_CURRENCY_NEEDED)}{r}/{_col(COL_AVG_GAIN)}{r},"
        f"0)"
    )

    # Style calculated columns as black formula cells
    for col in [
        COL_CURRENCY_TOTAL,
        COL_PULLS_TOTAL,
        COL_TOTAL_PULLS,
        COL_CURRENCY_NEEDED,
        COL_AVG_GAIN,
        COL_EST_DAYS,
    ]:
        cell = ws.cell(r, col)
        cell.font = FORMULA_FONT
        cell.border = THIN_BORDER
        cell.number_format = "0.00"

    # Style input columns as blue
    for col in [COL_NET_CURRENCY, COL_PULLS_NET]:
        cell = ws.cell(r, col)
        cell.font = INPUT_FONT
        cell.border = THIN_BORDER

    # Alternate row shading
    if r % 2 == 0:
        for col in range(1, 10):
            ws.cell(r, col).fill = ALT_FILL

    # Date cell
    ws.cell(r, COL_DATE).border = THIN_BORDER


def _load_or_create_workbook(xlsx_file, currency_name):
    if os.path.exists(xlsx_file):
        return load_workbook(xlsx_file)
    wb = Workbook()
    ws = wb.active
    ws.title = "Diary Log"
    ws.freeze_panes = "A2"
    _apply_header(ws, currency_name)
    return wb


def _find_today_row(ws, today):
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
        cell = row[0]
        if str(cell.value) == today:
            return cell.row
    return None


def _last_data_row(ws):
    for r in range(ws.max_row, 1, -1):
        if ws.cell(r, COL_DATE).value is not None:
            return r
    return 1  # only header exists


async def update_diary_xlsx(client, uid, config: GameConfig):
    logger = logging.getLogger(f"update_{config.name.lower()}_diary")
    os.makedirs("data", exist_ok=True)

    diary = await config.diary_fetcher(client, uid)
    day_data = diary.day_data
    today = now().strftime("%Y-%m-%d")
    currency_gain = getattr(day_data, config.currency_attr)

    wb = _load_or_create_workbook(config.xlsx_file, config.currency_name)
    ws = wb.active

    # Remove today's row if it already exists (re-run scenario)
    existing_today_row = _find_today_row(ws, today)
    if existing_today_row:
        ws.delete_rows(existing_today_row)

    new_row = _last_data_row(ws) + 1

    # Write date and the two user-editable input values
    ws.cell(new_row, COL_DATE).value = today
    ws.cell(new_row, COL_NET_CURRENCY).value = currency_gain
    ws.cell(new_row, COL_PULLS_NET).value = 0

    # Write all formula-driven columns
    _write_row_formulas(ws, new_row, config.pull_cost, config.five_star_pity)

    wb.save(config.xlsx_file)
    logger.info(f"{config.name} diary updated successfully (row {new_row}).")
    return {"Date": today, "Net Currency Gain": currency_gain, "Pulls Net Gain": 0}